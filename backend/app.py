import os
import json
import time
import hashlib
import threading
import traceback
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import Flask, request, jsonify, send_file, send_from_directory
import sqlite3
import openpyxl
from openpyxl import Workbook
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
import anthropic
import re

app = Flask(__name__, static_folder='../frontend', static_url_path='')

# ── Storage paths ──
# On Railway: mount a Volume at /data (Settings → Volumes → Mount path: /data)
# Locally: uses ./data relative to project root
DATA_DIR   = os.environ.get('RAILWAY_VOLUME_MOUNT_PATH',
                            os.path.join(os.path.dirname(__file__), '..', 'data'))
DB_PATH    = os.path.join(DATA_DIR, 'db.sqlite')
UPLOAD_DIR = os.path.join(DATA_DIR, 'uploads')
EXPORT_DIR = os.path.join(DATA_DIR, 'exports')

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(EXPORT_DIR, exist_ok=True)

anthropic_client = anthropic.Anthropic(api_key=os.environ.get('ANTHROPIC_API_KEY', ''))

# ─────────────────────────────────────────
# FIX 1: Fuzzy column name matching
# Handles any Excel column naming variation
# ─────────────────────────────────────────

COLUMN_ALIASES = {
    'name':     ['שם עורך דין \\ משרד', 'שם בית העסק', 'שם עורך דין', 'שם משרד', 'שם'],
    'site':     ['אתר בית', 'אתר', 'website', 'url'],
    'facebook': ['פייסבוק', 'עמוד פייסבוק', 'facebook'],
    'city':     ['ישוב', 'עיר', 'city'],
    'category': ['קטגוריה', 'תחום', 'category', 'התמחות'],
}

def find_column(headers, aliases_key):
    """Return the actual header string that matches the semantic field."""
    candidates = COLUMN_ALIASES.get(aliases_key, [])
    for candidate in candidates:
        for h in headers:
            if h and candidate in str(h):
                return str(h)
    return None

def extract_fields(data, col_headers):
    """Extract semantic fields from a raw row dict using fuzzy column matching."""
    col = {key: find_column(col_headers, key) for key in COLUMN_ALIASES}
    return {
        'name':     str(data.get(col['name'], '') or '').strip(),
        'site':     str(data.get(col['site'], '') or '').strip(),
        'facebook': str(data.get(col['facebook'], '') or '').strip(),
        'city':     str(data.get(col['city'], '') or '').strip(),
    }

# ─────────────────────────────────────────
# Database
# ─────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH, timeout=30, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")  # safe for concurrent threads
    return conn

def init_db():
    conn = get_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS jobs (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            filename       TEXT,
            status         TEXT DEFAULT 'pending',
            mode           TEXT DEFAULT 'full',
            total_rows     INTEGER DEFAULT 0,
            processed_rows INTEGER DEFAULT 0,
            yes_count      INTEGER DEFAULT 0,
            no_count       INTEGER DEFAULT 0,
            maybe_count    INTEGER DEFAULT 0,
            error_count    INTEGER DEFAULT 0,
            other_count    INTEGER DEFAULT 0,
            created_at     TEXT,
            completed_at   TEXT,
            config         TEXT,
            col_headers    TEXT
        );
        CREATE TABLE IF NOT EXISTS lawyer_rows (
            id                    INTEGER PRIMARY KEY AUTOINCREMENT,
            job_id                INTEGER,
            row_index             INTEGER,
            raw_data              TEXT,
            site_final            TEXT,
            site_status           TEXT,
            primary_area_1        TEXT,
            secondary_area_1      TEXT,
            secondary_area_2      TEXT,
            confidence            INTEGER,
            recommendation        TEXT,
            recommendation_reason TEXT,
            evidence_1            TEXT,
            evidence_2            TEXT,
            facebook_found        TEXT,
            checked_at            TEXT,
            status                TEXT DEFAULT 'pending',
            error                 TEXT
        );
        CREATE TABLE IF NOT EXISTS site_cache (
            url           TEXT PRIMARY KEY,
            site_final    TEXT,
            classification TEXT,
            crawl_success  INTEGER DEFAULT 0,
            content_hash  TEXT,
            last_checked  TEXT
        );
    ''')
    conn.commit()
    conn.close()

init_db()

# ─────────────────────────────────────────
# HTTP / Crawling
# ─────────────────────────────────────────

HTTP_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (compatible; LegalResearcher/1.0)',
    'Accept-Language': 'he,en;q=0.9',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
}
PRIORITY_KEYWORDS = ['תחומי', 'שירותים', 'practice', 'services',
                     'about', 'אודות', 'expertise', 'areas', 'תחום']

def normalize_url(url):
    if not url:
        return None
    url = str(url).strip().rstrip('/')
    # Reject emails — they look like URLs but aren't
    if '@' in url and not url.startswith('http'):
        return None
    if not url.startswith('http'):
        url = 'https://' + url
    # Reject known bad domains immediately
    from urllib.parse import urlparse as _up
    p = _up(url)
    if any(bad in p.netloc for bad in BAD_DOMAINS):
        return None
    return url

def fetch_url(url, timeout=12):
    """Try https, fall back to http. Returns (final_url, html) or (None, None)."""
    if not url:
        return None, None
    attempts = [url]
    if url.startswith('https://'):
        attempts.append(url.replace('https://', 'http://'))
    for u in attempts:
        try:
            r = requests.get(u, headers=HTTP_HEADERS, timeout=timeout, allow_redirects=True)
            if r.status_code == 200:
                return r.url, r.text
        except Exception:
            pass
    return None, None

def extract_clean_text(soup):
    for tag in soup.find_all(['nav', 'footer', 'script', 'style', 'header']):
        tag.decompose()
    texts = []
    for h in soup.find_all(['h1', 'h2', 'h3']):
        t = h.get_text(' ', strip=True)
        if t:
            texts.append(f"[HEADING] {t}")
    for li in soup.find_all('li'):
        t = li.get_text(' ', strip=True)
        if t and len(t) < 200:
            texts.append(f"[ITEM] {t}")
    for p in soup.find_all('p'):
        t = p.get_text(' ', strip=True)
        if t and len(t) > 30:
            texts.append(t)
    return '\n'.join(texts)[:8000]

def crawl_site(start_url, max_depth=2, max_pages=15):
    """Controlled crawl staying within same domain. Returns list of (url, text)."""
    base_domain = urlparse(start_url).netloc
    visited, queue, pages = set(), [(start_url, 0)], []

    while queue and len(pages) < max_pages:
        url, depth = queue.pop(0)
        if url in visited:
            continue
        visited.add(url)

        _, html = fetch_url(url)
        if not html:
            continue

        soup = BeautifulSoup(html, 'html.parser')
        text = extract_clean_text(soup)
        if text.strip():
            pages.append((url, text))

        if depth < max_depth:
            links = []
            for a in soup.find_all('a', href=True):
                full = urljoin(url, a['href'])
                p    = urlparse(full)
                if p.netloc == base_domain and full not in visited:
                    score = sum(1 for kw in PRIORITY_KEYWORDS
                                if kw.lower() in full.lower()
                                or kw.lower() in a.get_text().lower())
                    links.append((score, full))
            links.sort(reverse=True)
            for _, link in links[:20]:
                queue.append((link, depth + 1))

    return pages

# ─────────────────────────────────────────
# FIX 2 & 3: Claude calls with retry
# ─────────────────────────────────────────

def claude_with_retry(fn, retries=3):
    """Call fn() which makes an Anthropic API call. Retry with exponential backoff."""
    for attempt in range(retries):
        try:
            result = fn()
            if result is not None:
                return result
        except TypeError as e:
            # TypeError usually means bad input to API - log full traceback once
            import traceback as _tb
            if attempt == 0:
                print(f"  [Claude TypeError] {_tb.format_exc()[-300:]}")
            else:
                print(f"  [Claude retry {attempt+1}/{retries}] TypeError: {str(e)[:80]}")
            if attempt < retries - 1:
                time.sleep(1)
        except Exception as e:
            print(f"  [Claude retry {attempt+1}/{retries}] {str(e)[:100]}")
            if attempt < retries - 1:
                time.sleep(2 ** attempt)
    return None

BAD_DOMAINS = ['google', 'facebook', 'fb.com', 'anthropic', 'psakdin', 'din.co.il',
               'martindale', 'lawtech', 'b144', 'dun', 'linkedin',
               'yad2', 'bizportal', 'mako', 'walla', 'ynet', 'zap.co.il',
               'gov.il', 'court.gov', 'wikipedia', 'instagram', 'twitter']

def extract_all_text_from_response(response):
    """Extract all text content from an Anthropic API response including tool results."""
    texts = []
    for block in response.content:
        if hasattr(block, 'text') and block.text:
            texts.append(block.text)
        elif hasattr(block, 'content'):
            inner = block.content
            if isinstance(inner, list):
                for item in inner:
                    if hasattr(item, 'text') and item.text:
                        texts.append(item.text)
            elif isinstance(inner, str):
                texts.append(inner)
    return '\n'.join(texts)

def web_search_for_site(lawyer_name, city):
    """Search Google for the lawyer and return both a URL and search snippet text."""
    if not lawyer_name:
        return None, None
    query = f"{lawyer_name} עורך דין {city or ''}"

    def _call():
        response = anthropic_client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1000,
            tools=[{"type": "web_search_20250305", "name": "web_search"}],
            messages=[{"role": "user", "content":
                f"Search for this Israeli lawyer: {query}. "
                f"Return their official website URL and a brief description of their practice areas based on search results."}]
        )
        # Extract only clean text blocks (not tool_use blocks which contain dicts)
        clean_texts = []
        found_url = None
        for block in response.content:
            # Get text from text blocks
            if hasattr(block, 'type') and block.type == 'text' and hasattr(block, 'text'):
                clean_texts.append(block.text)
                # Extract URL from text
                if not found_url:
                    for url in re.findall(r'https?://[^\s\'"<>]+', block.text):
                        p = urlparse(url)
                        if p.netloc and not any(bad in p.netloc for bad in BAD_DOMAINS):
                            found_url = url
                            break
            # Get text from tool_result content (search results)
            elif hasattr(block, 'type') and block.type == 'tool_result':
                inner = block.content if hasattr(block, 'content') else None
                if isinstance(inner, list):
                    for item in inner:
                        if hasattr(item, 'type') and item.type == 'text' and hasattr(item, 'text'):
                            clean_texts.append(item.text)
                            if not found_url:
                                for url in re.findall(r'https?://[^\s\'"<>]+', item.text):
                                    p = urlparse(url)
                                    if p.netloc and not any(bad in p.netloc for bad in BAD_DOMAINS):
                                        found_url = url
                                        break
        
        snippet = '\n'.join(clean_texts).strip()
        return found_url, snippet if snippet else None

    try:
        result = _call()
        if result is None:
            return None, None
        return result
    except Exception as e:
        print(f"  [web_search error] {str(e)[:100]}")
        return None, None

def web_search_for_facebook(lawyer_name):
    if not lawyer_name:
        return None

    def _call():
        response = anthropic_client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=300,
            tools=[{"type": "web_search_20250305", "name": "web_search"}],
            messages=[{"role": "user", "content":
                f"Find the Facebook page URL for Israeli lawyer: {lawyer_name}. "
                f"Return ONLY the facebook.com URL if found, nothing else."}]
        )
        for block in response.content:
            if hasattr(block, 'text'):
                urls = re.findall(r'https?://(?:www\.)?facebook\.com/[^\s\'"<>]+', block.text)
                if urls:
                    return urls[0]
        return None

    return claude_with_retry(_call)

PRACTICE_AREAS = [
    "משפט אזרחי",
    "משפט מסחרי / דיני חברות",
    "דיני חוזים",
    "דיני נזיקין",
    "דיני בנקאות",
    "דיני מקרקעין (נדל\"ן)",
    "תכנון ובנייה",
    "משפט פלילי",
    "משפט מנהלי",
    "משפט חוקתי",
    "דיני עבודה",
    "דיני משפחה",
    "דיני ירושה וצוואות",
    "קניין רוחני",
    "דיני מיסים",
    "חדלות פירעון (פשיטת רגל ושיקום כלכלי)",
    "דיני תחרות (הגבלים עסקיים)",
    "משפט בינלאומי פרטי",
    "משפט בינלאומי פומבי",
    "משפט ימי",
    "משפט צבאי",
    "דיני הגירה ואזרחות",
    "דיני איכות הסביבה",
    "דיני תקשורת ומדיה",
    "דיני פרטיות והגנת מידע",
    "דיני מכרזים",
    "דיני ספורט",
    "דיני צרכנות",
    "אחר",
]
AREAS_LIST_STR = "\n".join(f"{i+1}. {a}" for i, a in enumerate(PRACTICE_AREAS))

def classify_practice_areas(text_corpus, lawyer_name, category_hint=''):
    if not text_corpus or not text_corpus.strip():
        return None

    hint_line = f"\n\nנתון נוסף מהמאגר: הקטגוריה הרשומה היא '{category_hint}'. השתמש בה כרמז לסיווג." if category_hint else ""
    prompt = f"""You are analyzing the website of an Israeli lawyer or law firm named "{lawyer_name}".

Website content:
{text_corpus[:5000]}

Classify this lawyer using ONLY the following 29 practice area categories (plus "אחר"):
{AREAS_LIST_STR}

Sub-areas that belong to each category (use these to help identify the right category):
- משפט אזרחי: סדר דין אזרחי, תביעות, סעדים זמניים, עיקולים, צווי מניעה, התיישנות
- משפט מסחרי / דיני חברות: ייסוד חברות, ממשל תאגידי, סכסוכי בעלי מניות, מיזוגים ורכישות, שותפויות
- דיני חוזים: חוזים, הפרת חוזים, פרשנות חוזים
- דיני נזיקין: רשלנות, רשלנות רפואית, תאונות דרכים, תאונות עבודה, ביטוח לאומי, לשון הרע, ביטוח
- דיני בנקאות: סכסוכי לקוח-בנק, הלוואות, ערבויות, משכנתאות, שטרות
- דיני מקרקעין (נדל"ן): עסקאות מכר, שכירות, בתים משותפים, התחדשות עירונית, תמ"א 38, ליקויי בנייה
- תכנון ובנייה: היתרי בנייה, התנגדויות, עבירות תכנוניות, היטל השבחה
- משפט פלילי: עבירות רכוש, אלימות, סמים, מרמה, עבירות כלכליות, הלבנת הון, העלמות מס
- משפט מנהלי: עתירות מנהליות, תקיפת החלטות רשות, רישוי, ביקורת שיפוטית
- משפט חוקתי: זכויות יסוד, עתירות לבג"ץ, ביקורת שיפוטית על חקיקה
- דיני עבודה: פיטורים, זכויות סוציאליות, אפליה בעבודה, הטרדה מינית, הסכמים קיבוציים
- דיני משפחה: גירושין, מזונות, משמורת, חלוקת רכוש, הסכמי ממון, ידועים בציבור
- דיני ירושה וצוואות: צוואות, ירושה, צו ירושה, סכסוכי יורשים, אפוטרופסות, ייפוי כוח מתמשך
- קניין רוחני: זכויות יוצרים, סימני מסחר, פטנטים, סודות מסחריים
- דיני מיסים: מס הכנסה, מע"מ, מיסוי מקרקעין, השגות מס, קריפטו, גילוי מרצון
- חדלות פירעון: פשיטת רגל, הסדרי חוב, הוצאה לפועל, גביה, הגנה מפני נושים
- דיני תחרות: הגבלים עסקיים, מונופולין, הסדרים כובלים, תביעות נגזרות
- משפט בינלאומי פרטי: סמכות שיפוט בינלאומית, הכרה בפסקי דין זרים
- משפט בינלאומי פומבי: דיני אמנות, אחריות מדינתית, דיני לחימה
- משפט ימי: הובלה ימית, מטען, כלי שיט, תאונות ימיות
- משפט צבאי: דין משמעתי, בתי דין צבאיים, זכויות חיילים
- דיני הגירה ואזרחות: אשרות, מעמד, התאזרחות, איחוד משפחות, דרכון ישראלי
- דיני איכות הסביבה: רגולציה סביבתית, מפגעים, אחריות סביבתית
- דיני תקשורת ומדיה: רגולציית שידורים, דיני אינטרנט, נגישות אתרים
- דיני פרטיות והגנת מידע: חוק הגנת הפרטיות, מאגרי מידע, אבטחת מידע
- דיני מכרזים: מכרזים ציבוריים, השגות מכרזים, פסילת הצעות
- דיני ספורט: חוזי שחקנים, מוסדות שיפוט בספורט, משמעת בספורט
- דיני צרכנות: הטעיה, ביטול עסקה, תביעות ייצוגיות, אחריות לצרכן
- אחר: כל תחום שאינו נופל באף קטגוריה לעיל

Return ONLY valid JSON, no explanation:
{{
  "primary_practice_areas": ["exact name from list"],
  "secondary_practice_areas": ["exact name from list"],
  "confidence": 75,
  "evidence": ["short Hebrew snippet from the text"]
}}

Rules:
- primary_practice_areas: exactly 1 item — the single most dominant area on the site
- secondary_practice_areas: 0-2 items maximum — only if clearly present
- ONLY use exact category names from the numbered list
- Use "אחר" only if nothing fits
- confidence: 0-100{hint_line}"""
    def _call():
        response = anthropic_client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=800,
            messages=[{"role": "user", "content": prompt}]
        )
        block = response.content[0]
        text = block.text if hasattr(block, 'text') else (block.get('text','') if isinstance(block, dict) else '')
        # Find the outermost JSON object robustly
        start = text.find('{')
        if start == -1:
            return None
        depth = 0
        end = -1
        for i, ch in enumerate(text[start:], start):
            if ch == '{': depth += 1
            elif ch == '}':
                depth -= 1
                if depth == 0:
                    end = i + 1
                    break
        if end == -1:
            return None
        try:
            result = json.loads(text[start:end])
            if result.get('primary_practice_areas'):
                return result
        except (json.JSONDecodeError, TypeError):
            pass
        return None

    return claude_with_retry(_call)

# ─────────────────────────────────────────
# Business rules
# ─────────────────────────────────────────
def apply_business_rules(classification, config):
    if not classification or not classification.get('primary_practice_areas'):
        return "MAYBE", "לא ניתן לסווג — לא נמצא תוכן אתר"

    target   = [a.strip().lower() for a in config.get('target_areas', [])   if a.strip()]
    excluded = [a.strip().lower() for a in config.get('excluded_areas', []) if a.strip()]
    all_areas = [a.lower() for a in
                 classification.get('primary_practice_areas', []) +
                 classification.get('secondary_practice_areas', [])]

    for area in all_areas:
        for ex in excluded:
            if ex in area:
                return "NO", f"תחום מוחרג: {area}"

    for area in all_areas:
        for tgt in target:
            if tgt in area:
                return "YES", f"תחום מועדף: {area}"

    if not target:
        return "MAYBE", "סווג בהצלחה"

    return "MAYBE", "לא נמצא התאמה לתחומים מועדפים"

# ─────────────────────────────────────────
# Core row processor
# ─────────────────────────────────────────

def classify_from_name(lawyer_name):
    """Extract practice areas from the lawyer's name/title when no site is available."""
    prompt = f"""An Israeli lawyer is listed with this name/title: "{lawyer_name}"

Sometimes Israeli lawyer names include their practice areas (e.g. "עו"ד גלינה פסחוב - תעבורה, פלילי, מנהלי").

Choose their practice areas ONLY from this fixed list:
{AREAS_LIST_STR}

If you can identify practice areas from the name/title, return JSON:
{{
  "primary_practice_areas": ["exact name from list"],
  "secondary_practice_areas": [],
  "confidence": 60,
  "evidence": ["extracted from name: {lawyer_name}"]
}}

If the name gives NO information about practice areas, return null (nothing).
Return ONLY valid JSON or the word null."""

    def _call():
        response = anthropic_client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=300,
            messages=[{{"role": "user", "content": prompt}}]
        )
        text = response.content[0].text.strip()
        if text.lower() == 'null' or not text or '{{' not in text:
            return None
        match = re.search(r'\{{.*\}}', text, re.DOTALL)
        if match:
            result = json.loads(match.group())
            if result.get('primary_practice_areas'):
                return result
        return None

    return claude_with_retry(_call)


def process_row(job_id, row_id, raw_data, config, col_headers):
    conn = get_db()
    try:
        data   = json.loads(raw_data)
        fields = extract_fields(data, col_headers)

        lawyer_name    = fields['name']
        site_input     = fields['site']
        facebook_input = fields['facebook']
        city           = fields['city']
        # Get category hint from source data (e.g. "עורכי דין - דיני משפחה")
        category_hint  = ''
        cat_col = find_column(col_headers, 'category')
        if cat_col and cat_col in data:
            category_hint = str(data[cat_col] or '').strip()

        print(f"  → [{row_id}] {lawyer_name or '(no name)'} | site: {site_input or 'none'}")

        # ── Step 2: Resolve website ──
        site_final = site_status = None

        if site_input:
            url = normalize_url(site_input)
            if url:
                final_url, _ = fetch_url(url)
                if final_url:
                    site_final = final_url
                else:
                    site_status = "FETCH_FAILED"

        # Always try web search as fallback
        search_snippet = None
        if not site_final:
            found_url, search_snippet = web_search_for_site(lawyer_name, city)
            if found_url:
                final_url, _ = fetch_url(found_url)
                if final_url:
                    site_final  = final_url
                    site_status = "FOUND_VIA_SEARCH"
                else:
                    site_status = "SEARCH_FOUND_BUT_FETCH_FAILED"

        if not site_final:
            site_status = "NO_SITE"

        # ── Step 2b: Classify from name if no site found ──
        name_classification = None
        if not site_final and lawyer_name and lawyer_name.strip():
            name_classification = classify_from_name(lawyer_name)

        # ── Step 3: Cache check ──
        classification = None
        if site_final:
            cache = conn.execute(
                'SELECT * FROM site_cache WHERE url=?', (site_final,)).fetchone()
            if cache:
                age = datetime.now() - datetime.fromisoformat(cache['last_checked'])
                # FIX: only use cache if crawl was actually successful
                if age < timedelta(days=7) and cache['crawl_success'] and cache['classification']:
                    classification = json.loads(cache['classification'])
                    site_status    = "CACHED"
                    print(f"    ↩ Cache hit")

        # ── Steps 4–5: Crawl + classify ──
        if site_final and classification is None:
            pages  = crawl_site(site_final)
            corpus = '\n\n'.join(text for _, text in pages)
            # Also append search snippet to corpus for better classification
            if search_snippet:
                corpus = search_snippet + '\n\n' + corpus
            print(f"    ↳ Crawled {len(pages)} pages, {len(corpus)} chars")

            classification = classify_practice_areas(corpus, lawyer_name, category_hint)
            site_status    = site_status or ("CRAWLED" if pages else "NO_CONTENT")
        
        # ── If no site but have search snippet, classify from that ──
        elif not site_final and search_snippet and classification is None:
            print(f"    ↳ Classifying from search snippet ({len(search_snippet)} chars)")
            classification = classify_practice_areas(search_snippet, lawyer_name, category_hint)
            if classification:
                site_status = "CLASSIFIED_FROM_SEARCH"
                print(f"    ✓ {classification.get('primary_practice_areas')}")
            else:
                print(f"    ✗ Classification empty")

        # ── Step 6: Business rules ──
        # Use name-based classification as fallback
        if not classification and name_classification:
            classification = name_classification
        recommendation, reason = apply_business_rules(classification, config)

        # ── Save ──
        facebook_found = facebook_input
        primary    = classification.get('primary_practice_areas', []) if classification else []
        secondary  = classification.get('secondary_practice_areas', []) if classification else []
        evidence   = classification.get('evidence', [])                 if classification else []
        confidence = classification.get('confidence', 0)               if classification else 0

        conn.execute('''
            UPDATE lawyer_rows SET
              site_final=?, site_status=?,
              primary_area_1=?, secondary_area_1=?, secondary_area_2=?,
              confidence=?,
              recommendation=?, recommendation_reason=?,
              evidence_1=?, evidence_2=?,
              facebook_found=?, checked_at=?,
              status='done', error=NULL
            WHERE id=?''', (
            site_final, site_status,
            primary[0]    if len(primary) > 0 else None,
            secondary[0]  if len(secondary) > 0 else None,
            secondary[1]  if len(secondary) > 1 else None,
            confidence,
            recommendation, reason,
            evidence[0] if len(evidence) > 0 else None,
            evidence[1] if len(evidence) > 1 else None,
            facebook_found, datetime.now().isoformat(),
            row_id
        ))
        conn.commit()
        return recommendation

    except Exception:
        tb = traceback.format_exc()
        print(f"  ✗ Row {row_id} failed:\n{tb[:300]}")
        try:
            conn.execute(
                "UPDATE lawyer_rows SET status='error', error=? WHERE id=?",
                (tb[:1000], row_id))
            conn.commit()
        except Exception:
            pass
        return 'ERROR'

    finally:
        try:
            done = conn.execute(
                "SELECT COUNT(*) FROM lawyer_rows "
                "WHERE job_id=? AND status IN ('done','error')",
                (job_id,)).fetchone()[0]
            conn.execute(
                "UPDATE jobs SET processed_rows=? WHERE id=?", (done, job_id))
            conn.commit()
        except Exception:
            pass
        conn.close()

# ─────────────────────────────────────────
# FIX 2: Job runner — protected loop + parallel workers
# ─────────────────────────────────────────

MAX_WORKERS = 3

def run_job(job_id):
    conn = get_db()
    try:
        conn.execute("UPDATE jobs SET status='running' WHERE id=?", (job_id,))
        conn.commit()

        job         = conn.execute("SELECT * FROM jobs WHERE id=?", (job_id,)).fetchone()
        config      = json.loads(job['config'])      if job['config']      else {}
        col_headers = json.loads(job['col_headers']) if job['col_headers'] else []
        rows        = conn.execute(
            "SELECT * FROM lawyer_rows WHERE job_id=? AND status='pending' ORDER BY row_index",
            (job_id,)).fetchall()
        conn.close()
        conn = None

        print(f"[Job {job_id}] {len(rows)} pending rows | mode={job['mode']} | workers={MAX_WORKERS}")

        yes_c = no_c = maybe_c = err_c = other_c = 0

        # FIX: ThreadPoolExecutor with proper exception capture per row
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {
                executor.submit(
                    process_row, job_id, r['id'], r['raw_data'], config, col_headers
                ): r['id']
                for r in rows
            }
            for future in as_completed(futures):
                row_id = futures[future]
                try:
                    result = future.result()
                    if result == 'YES':     yes_c   += 1
                    elif result == 'NO':    no_c    += 1
                    elif result == 'MAYBE': maybe_c += 1
                    else:                  err_c   += 1
                except Exception as e:
                    print(f"[Job {job_id}] Row {row_id} future exception: {e}")
                    err_c += 1

        conn = get_db()
        # Count 'אחר' classifications
        other_c = conn.execute(
            "SELECT COUNT(*) FROM lawyer_rows WHERE job_id=? AND primary_area_1='אחר'",
            (job_id,)).fetchone()[0]

        conn.execute(
            '''UPDATE jobs SET status='completed', completed_at=?,
               yes_count=?, no_count=?, maybe_count=?, error_count=?, other_count=?
               WHERE id=?''',
            (datetime.now().isoformat(), yes_c, no_c, maybe_c, err_c, other_c, job_id))
        conn.commit()
        conn.close()
        print(f"[Job {job_id}] Complete ✓  YES:{yes_c} NO:{no_c} MAYBE:{maybe_c} ERR:{err_c}")

    except Exception:
        tb = traceback.format_exc()
        print(f"[Job {job_id}] FATAL thread crash:\n{tb}")
        try:
            if conn:
                conn.close()
            conn = get_db()
            conn.execute(
                "UPDATE jobs SET status='error', completed_at=? WHERE id=?",
                (datetime.now().isoformat(), job_id))
            conn.commit()
            conn.close()
        except Exception:
            pass

# ─────────────────────────────────────────
# API Routes
# ─────────────────────────────────────────

@app.route('/')
def index():
    return send_from_directory('../frontend', 'index.html')

@app.route('/api/health')
def health():
    return jsonify({'status': 'ok', 'time': datetime.now().isoformat()})

@app.route('/api/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    f          = request.files['file']
    config_str = request.form.get('config', '{}')
    mode       = request.form.get('mode', 'full')  # full / sample_10 / sample_50 / sample_100
    config     = json.loads(config_str)

    path = os.path.join(UPLOAD_DIR, f"{int(time.time())}_{f.filename}")
    f.save(path)

    wb      = openpyxl.load_workbook(path)
    ws      = wb.active
    first_row = [str(c.value) if c.value is not None else '' for c in ws[1]]

    # Auto-detect if first row is a header or data
    # If the first cell looks like a name (not a known column label), treat as data
    KNOWN_HEADERS = ['שם', 'name', 'עורך', 'משרד', 'טלפון', 'אימייל', 'email', 'אתר', 'site', 'עיר', 'ישוב', 'city']
    first_cell = first_row[0].strip().lower() if first_row else ''
    has_header = any(kw in first_cell for kw in KNOWN_HEADERS)

    if has_header:
        headers  = first_row
        data_start = 2
    else:
        # No header row — assign positional column names based on content sniffing
        # Detect which column has URLs and which has names
        headers = []
        for j, val in enumerate(first_row):
            v = str(val).strip() if val else ''
            if ('@' not in v) and ('http' in v or '.co.il' in v or '.com' in v):
                headers.append('אתר בית')
            elif '@' in v:
                headers.append('אימייל')
            elif v.replace('-','').replace('+','').replace(' ','').isdigit() and len(v) > 5:
                headers.append('טלפון')
            elif j == 0:
                headers.append('שם עורך דין')
            else:
                headers.append(f'col_{j}')
        data_start = 1

    conn = get_db()
    c    = conn.cursor()
    c.execute(
        "INSERT INTO jobs (filename, status, mode, total_rows, created_at, config, col_headers) "
        "VALUES (?,?,?,?,?,?,?)",
        (f.filename, 'pending', mode, 0,
         datetime.now().isoformat(), json.dumps(config), json.dumps(headers)))
    job_id = c.lastrowid

    # Parse all rows
    all_raw = []
    for i, row in enumerate(ws.iter_rows(min_row=data_start, values_only=True)):
        if not any(v for v in row if v is not None):
            continue
        row_data = {}
        for j in range(min(len(headers), len(row))):
            key = headers[j]
            val = row[j]
            # If multiple cols mapped to same name, keep first non-empty
            if key not in row_data or (row_data[key] in (None, '')):
                row_data[key] = val
            # Also check each cell for URL pattern if no site col found yet
            if 'אתר בית' not in row_data or not row_data.get('אתר בית'):
                v = str(val).strip() if val else ''
                # Only pick up URLs, not emails
                if '@' not in v and ('http' in v or '.co.il' in v or '.com' in v) and len(v) > 5:
                    row_data['אתר בית'] = val
        all_raw.append((i, json.dumps(row_data, ensure_ascii=False, default=str)))

    # Determine which rows are 'pending' vs 'skipped' based on mode
    SAMPLE_LIMITS = {'sample_10': 10, 'sample_50': 50, 'sample_100': 100}

    if mode in SAMPLE_LIMITS:
        limit    = SAMPLE_LIMITS[mode]
        with_site    = []
        without_site = []
        for i, rd in all_raw:
            row_obj  = json.loads(rd)
            site_val = ''
            for v in row_obj.values():
                s = str(v or '').strip()
                if 'http' in s or '.co.il' in s or '.com' in s:
                    site_val = s
                    break
            (with_site if site_val else without_site).append((i, rd))

        # Smart sample: prefer rows with site, pad with rows without if needed
        chosen = with_site[:limit]
        if len(chosen) < limit:
            chosen += without_site[:limit - len(chosen)]
        chosen_set = {i for i, _ in chosen}
    else:
        chosen_set = {i for i, _ in all_raw}

    for i, rd in all_raw:
        status = 'pending' if i in chosen_set else 'skipped'
        c.execute(
            "INSERT INTO lawyer_rows (job_id, row_index, raw_data, status) VALUES (?,?,?,?)",
            (job_id, i, rd, status))

    total_pending = len(chosen_set)
    c.execute("UPDATE jobs SET total_rows=? WHERE id=?", (total_pending, job_id))
    conn.commit()
    conn.close()

    t = threading.Thread(target=run_job, args=(job_id,), daemon=True)
    t.start()

    return jsonify({
        'job_id':       job_id,
        'total_rows':   total_pending,
        'skipped_rows': len(all_raw) - total_pending,
        'mode':         mode,
    })

@app.route('/api/jobs/<int:job_id>/resume', methods=['POST'])
def resume_job(job_id):
    """Resume processing any pending/error rows in a stopped job."""
    conn = get_db()
    job  = conn.execute("SELECT * FROM jobs WHERE id=?", (job_id,)).fetchone()
    if not job:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    pending = conn.execute(
        "SELECT COUNT(*) FROM lawyer_rows WHERE job_id=? AND status='pending'",
        (job_id,)).fetchone()[0]
    conn.execute(
        "UPDATE jobs SET status='running', completed_at=NULL WHERE id=?", (job_id,))
    conn.commit()
    conn.close()

    t = threading.Thread(target=run_job, args=(job_id,), daemon=True)
    t.start()
    return jsonify({'ok': True, 'resuming_rows': pending})

@app.route('/api/jobs', methods=['GET'])
def list_jobs():
    conn = get_db()
    jobs = conn.execute("SELECT * FROM jobs ORDER BY id DESC LIMIT 30").fetchall()
    conn.close()
    return jsonify([dict(j) for j in jobs])

@app.route('/api/jobs/<int:job_id>', methods=['GET'])
def get_job(job_id):
    conn = get_db()
    job  = conn.execute("SELECT * FROM jobs WHERE id=?", (job_id,)).fetchone()
    conn.close()
    if not job:
        return jsonify({'error': 'Not found'}), 404
    return jsonify(dict(job))

@app.route('/api/jobs/<int:job_id>/rows', methods=['GET'])
def get_rows(job_id):
    page       = int(request.args.get('page', 1))
    per_page   = int(request.args.get('per_page', 20))
    filter_rec = request.args.get('recommendation', '')
    search     = request.args.get('search', '')
    offset     = (page - 1) * per_page

    conn   = get_db()
    where  = "WHERE job_id=? AND status != 'skipped'"
    params = [job_id]
    if filter_rec:
        where  += " AND recommendation=?"
        params.append(filter_rec)
    if search:
        where  += " AND (raw_data LIKE ? OR site_final LIKE ?)"
        params += [f'%{search}%', f'%{search}%']

    total = conn.execute(
        f"SELECT COUNT(*) FROM lawyer_rows {where}", params).fetchone()[0]
    rows  = conn.execute(
        f"SELECT * FROM lawyer_rows {where} ORDER BY row_index LIMIT ? OFFSET ?",
        params + [per_page, offset]).fetchall()
    conn.close()

    result = []
    for r in rows:
        d = dict(r)
        d['raw_data'] = json.loads(d['raw_data']) if d['raw_data'] else {}
        result.append(d)

    return jsonify({'rows': result, 'total': total, 'page': page, 'per_page': per_page})

@app.route('/api/jobs/<int:job_id>/export', methods=['GET'])
def export_job(job_id):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM lawyer_rows WHERE job_id=? ORDER BY row_index",
        (job_id,)).fetchall()
    conn.close()

    if not rows:
        return jsonify({'error': 'No rows'}), 404

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    first_raw    = json.loads(rows[0]['raw_data'])
    orig_headers = list(first_raw.keys())
    extra_headers = ['אתר סופי', 'סטטוס אתר',
                     'תחום עיקרי', 'תחום משני 1', 'תחום משני 2',
                     'תאריך בדיקה', 'שגיאה']
    ws.append(orig_headers + extra_headers)

    for row in rows:
        raw  = json.loads(row['raw_data']) if row['raw_data'] else {}
        vals = [raw.get(h, '') for h in orig_headers] + [
            row['site_final']   or '',
            row['site_status']  or '',
            row['primary_area_1']   or '',
            row['secondary_area_1'] or '',
            row['secondary_area_2'] or '',
            row['checked_at']   or '',
            row['error']        or '',
        ]
        ws.append(vals)

    out = os.path.join(EXPORT_DIR, f"enriched_{job_id}_{int(time.time())}.xlsx")
    wb.save(out)
    return send_file(out, as_attachment=True,
                     download_name=f"enriched_lawyers_{job_id}.xlsx")

@app.route('/api/cache/clear', methods=['POST'])
def clear_cache():
    conn = get_db()
    conn.execute("DELETE FROM site_cache")
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/cache/stats', methods=['GET'])
def cache_stats():
    conn  = get_db()
    total = conn.execute("SELECT COUNT(*) FROM site_cache").fetchone()[0]
    good  = conn.execute(
        "SELECT COUNT(*) FROM site_cache WHERE crawl_success=1").fetchone()[0]
    conn.close()
    return jsonify({'total': total, 'successful': good, 'empty': total - good})

if __name__ == '__main__':
    port  = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') == 'development'
    app.run(host='0.0.0.0', port=port, debug=debug, threaded=True)
